#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Parse 比赛数据.xlsx with composite anomaly scoring.
Generates:
  - data/dashboard-data.json  (for frontend)
  - data/analysis-debug.json  (for manual verification)
"""

import json
import math
import os
import sys
from collections import defaultdict
from copy import deepcopy

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Config ───────────────────────────────────────────────────────
LATEST_QUARTER = "2025Q1"
PREV_QUARTER = "2024Q4"       # for QoQ growth
LATEST_MONTHS = ["202501", "202502", "202503"]
PREV_MONTHS = ["202410", "202411", "202412"]

# Composite score weights
BU_WEIGHTS = {"规模": 0.40, "占比": 0.25, "增长率": 0.20, "积压": 0.15}
PL_WEIGHTS = {"规模": 0.35, "占比": 0.20, "增长率": 0.20, "高风险物料": 0.25}
MAT_WEIGHTS = {"金额": 0.40, "库龄": 0.35, "超360天": 0.25}

# ── Helpers ──────────────────────────────────────────────────────
def minmax_normalize(values):
    """Min-max normalize a list of numbers. Returns list of scores 0-1."""
    if not values:
        return []
    vmin = min(values)
    vmax = max(values)
    if vmax == vmin:
        return [0.0] * len(values)
    return [(v - vmin) / (vmax - vmin) for v in values]


def safe_float(v, default=0.0):
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def is_valid_number(v):
    """Check if value is a valid finite number."""
    if v is None:
        return False
    try:
        f = float(v)
        return math.isfinite(f)
    except (ValueError, TypeError):
        return False


# ── File loading ─────────────────────────────────────────────────
def load_workbook():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    for candidate in [
        os.path.join(os.path.dirname(project_dir), "比赛数据.xlsx"),
        os.path.join(project_dir, "比赛数据.xlsx"),
        os.path.join(os.getcwd(), "比赛数据.xlsx"),
    ]:
        if os.path.exists(candidate):
            return openpyxl.load_workbook(candidate, data_only=True), candidate
    raise FileNotFoundError("Cannot find 比赛数据.xlsx")


# ── Sheet parsers ────────────────────────────────────────────────
def parse_sheet1(ws):
    """Sheet 1: 库存成本趋势 - monthly cost by BU & PL."""
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    rows = []
    null_count = 0
    non_numeric = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            null_count += 1
            continue
        # Validate numeric columns (indices 4-11)
        for j in [4, 5, 6, 7]:
            if not is_valid_number(row[j]):
                non_numeric += 1
        rows.append({
            "月份": str(row[0]),
            "季度": str(row[1]) if row[1] else "",
            "业务单元": str(row[2]) if row[2] else "",
            "产品线": str(row[3]) if row[3] else "",
            "总库存成本": safe_float(row[4]),
            "原材料成本": safe_float(row[5]),
            "半成品成本": safe_float(row[6]),
            "产成品成本": safe_float(row[7]),
        })
    return rows, {
        "sheet_name": "库存成本趋势",
        "headers": headers,
        "total_rows": ws.max_row,
        "data_rows": len(rows),
        "columns": ws.max_column,
        "null_rows": null_count,
        "non_numeric_fields": non_numeric,
    }


def parse_sheet2(ws):
    """Sheet 2: 物料库存明细 - material inventory with aging."""
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    rows = []
    null_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            null_count += 1
            continue
        rows.append({
            "入库日期": str(row[0]) if row[0] else "",
            "业务单元": str(row[1]) if row[1] else "",
            "产品线": str(row[2]) if row[2] else "",
            "原材料成本": safe_float(row[3]),
            "物料编码": str(row[4]) if row[4] else "",
            "物料名称": str(row[5]) if row[5] else "",
            "当前库存量": safe_float(row[6]),
            "库龄天数": safe_float(row[7]),
            "库龄分段": str(row[8]) if row[8] else "",
            "批次号": str(row[9]) if row[9] else "",
        })
    return rows, {
        "sheet_name": "物料库存明细",
        "headers": headers[:10],
        "total_rows": ws.max_row,
        "data_rows": len(rows),
        "columns": ws.max_column,
        "null_rows": null_count,
    }


def parse_sheet3(ws):
    """Sheet 3: 库存成本结构 - quarterly cost structure."""
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "季度": str(row[0]) if row[0] else "",
            "成本构成": str(row[1]) if row[1] else "",
            "值": safe_float(row[2]),
        })
    quarters = sorted(set(r["季度"] for r in rows))
    cost_types = sorted(set(r["成本构成"] for r in rows))
    return rows, {
        "sheet_name": "库存成本结构",
        "headers": headers[:3],
        "total_rows": ws.max_row,
        "data_rows": len(rows),
        "columns": ws.max_column,
        "quarters_found": quarters,
        "cost_types_found": cost_types,
    }


# ── Quarterly cost structure (from Sheet 3) ──────────────────────
def build_structure_summary(sheet3_rows):
    """Pivot Sheet 3 data into quarterly structure."""
    qmap = defaultdict(lambda: {"原材料成本": 0, "半成品成本": 0, "产成品成本": 0})
    for r in sheet3_rows:
        ct = r["成本构成"]
        if ct in qmap[r["季度"]]:
            qmap[r["季度"]][ct] = r["值"]
    result = []
    for q in sorted(qmap.keys()):
        d = qmap[q]
        d["季度"] = q
        d["总库存成本"] = d["原材料成本"] + d["半成品成本"] + d["产成品成本"]
        d["原材料成本占比"] = round(d["原材料成本"] / d["总库存成本"] * 100, 2) if d["总库存成本"] > 0 else 0
        result.append(d)
    return result


# ── Aggregate Sheet 1 by quarter/BU/PL ───────────────────────────
def aggregate_by_quarter(sheet1_rows):
    """Aggregate monthly rows into quarterly totals."""
    qmap = defaultdict(lambda: {"总库存成本": 0, "原材料成本": 0, "半成品成本": 0, "产成品成本": 0})
    for r in sheet1_rows:
        q = r["季度"]
        qmap[q]["总库存成本"] += r["总库存成本"]
        qmap[q]["原材料成本"] += r["原材料成本"]
        qmap[q]["半成品成本"] += r["半成品成本"]
        qmap[q]["产成品成本"] += r["产成品成本"]
    result = []
    for q in sorted(qmap.keys()):
        d = qmap[q]
        d["季度"] = q
        d["原材料成本占比"] = round(d["原材料成本"] / d["总库存成本"] * 100, 2) if d["总库存成本"] > 0 else 0
        result.append(d)
    return result


def get_quarterly_by_bu(sheet1_rows):
    """Get quarterly aggregates per business unit."""
    bu_q = defaultdict(lambda: defaultdict(lambda: {"总库存成本": 0, "原材料成本": 0}))
    for r in sheet1_rows:
        bu = r["业务单元"]
        q = r["季度"]
        bu_q[bu][q]["总库存成本"] += r["总库存成本"]
        bu_q[bu][q]["原材料成本"] += r["原材料成本"]
    return bu_q


def get_quarterly_by_pl(sheet1_rows):
    """Get quarterly aggregates per product line."""
    pl_q = defaultdict(lambda: defaultdict(lambda: {"总库存成本": 0, "原材料成本": 0}))
    pl_bu = {}
    for r in sheet1_rows:
        pl = r["产品线"]
        q = r["季度"]
        pl_q[pl][q]["总库存成本"] += r["总库存成本"]
        pl_q[pl][q]["原材料成本"] += r["原材料成本"]
        if pl not in pl_bu:
            pl_bu[pl] = r["业务单元"]
    return pl_q, pl_bu


# ── Material analysis ────────────────────────────────────────────
def analyze_materials(sheet2_rows):
    """Full material analysis with risk scoring."""
    over_360 = [m for m in sheet2_rows if m["库龄天数"] > 360]
    stuck_amount = sum(m["原材料成本"] for m in over_360)

    # Aging distribution
    aging_map = defaultdict(lambda: {"物料数": 0, "金额合计": 0.0})
    for m in sheet2_rows:
        seg = m["库龄分段"]
        aging_map[seg]["物料数"] += 1
        aging_map[seg]["金额合计"] += m["原材料成本"]

    seg_order = ["30天以内", "31-90天", "91-180天", "181-360天", "360天以上"]
    aging_dist = []
    for seg in seg_order:
        if seg in aging_map:
            d = aging_map[seg]
            d["库龄分段"] = seg
            d["金额合计"] = round(d["金额合计"], 2)
            aging_dist.append(d)

    # Risk value: aging_days * amount / 10000
    for m in sheet2_rows:
        m["风险值"] = round(m["库龄天数"] * m["原材料成本"] / 10000, 2)

    # Composite material risk score (min-max normalized)
    amounts = [m["原材料成本"] for m in sheet2_rows]
    ages = [m["库龄天数"] for m in sheet2_rows]
    n_amounts = minmax_normalize(amounts)
    n_ages = minmax_normalize(ages)

    for i, m in enumerate(sheet2_rows):
        over_flag = 1.0 if m["库龄天数"] > 360 else 0.0
        m["风险得分"] = round(
            n_amounts[i] * MAT_WEIGHTS["金额"]
            + n_ages[i] * MAT_WEIGHTS["库龄"]
            + over_flag * MAT_WEIGHTS["超360天"],
            4
        )
        m["是否超360天"] = m["库龄天数"] > 360

    # TOP10 by risk score
    top_risk = sorted(sheet2_rows, key=lambda x: (-x["风险得分"], -x["风险值"]))

    # Group by material name
    mat_groups = defaultdict(lambda: {"总成本": 0.0, "总库存量": 0.0, "最大库龄": 0.0, "最大风险得分": 0.0, "物料编码": "", "批次数": 0})
    for m in sheet2_rows:
        nm = m["物料名称"]
        mat_groups[nm]["总成本"] += m["原材料成本"]
        mat_groups[nm]["总库存量"] += m["当前库存量"]
        mat_groups[nm]["最大库龄"] = max(mat_groups[nm]["最大库龄"], m["库龄天数"])
        mat_groups[nm]["最大风险得分"] = max(mat_groups[nm]["最大风险得分"], m["风险得分"])
        mat_groups[nm]["物料编码"] = m["物料编码"]
        mat_groups[nm]["批次数"] += 1

    mat_summary = []
    for nm, d in mat_groups.items():
        d["物料名称"] = nm
        d["总成本"] = round(d["总成本"], 2)
        d["总库存量"] = round(d["总库存量"], 2)
        d["最大库龄"] = round(d["最大库龄"], 2)
        mat_summary.append(d)
    mat_summary.sort(key=lambda x: -x["最大风险得分"])

    return {
        "物料明细": sheet2_rows,
        "超过360天物料": over_360,
        "超过360天物料数量": len(over_360),
        "积压金额_360天以上": round(stuck_amount, 2),
        "库龄分布": aging_dist,
        "高风险物料TOP10": top_risk[:10],
        "物料汇总": mat_summary,
    }


# ── Composite scoring: Business Units ────────────────────────────
def score_business_units(sheet1_rows, sheet2_rows, sheet3_structure):
    """Composite anomaly scoring for business units."""
    bu_q = get_quarterly_by_bu(sheet1_rows)

    # Material stuck amounts by BU
    bu_stuck = defaultdict(float)
    for m in sheet2_rows:
        if m["库龄天数"] > 360:
            bu_stuck[m["业务单元"]] += m["原材料成本"]

    # Collect raw metrics for latest quarter
    bu_names = sorted(bu_q.keys())
    raw_metrics = []
    for bu in bu_names:
        latest = bu_q[bu].get(LATEST_QUARTER, {"总库存成本": 0, "原材料成本": 0})
        prev = bu_q[bu].get(PREV_QUARTER, {"总库存成本": 0, "原材料成本": 0})
        total = latest["总库存成本"]
        raw = latest["原材料成本"]
        ratio = round(raw / total * 100, 2) if total > 0 else 0
        growth = round((raw - prev["原材料成本"]) / prev["原材料成本"] * 100, 2) if prev["原材料成本"] > 0 else 0
        stuck = round(bu_stuck.get(bu, 0) / 10000, 2)
        raw_metrics.append({
            "业务单元": bu,
            "最新期间总库存成本": round(total, 2),
            "最新期间原材料成本": round(raw, 2),
            "最新期间原材料成本占比": ratio,
            "期初原材料成本": round(prev["原材料成本"], 2),   # PREV_QUARTER
            "期末原材料成本": round(raw, 2),
            "原材料成本增长率": growth,
            "360天以上积压金额_万元": stuck,
        })

    # Min-max normalize
    scales = [m["最新期间原材料成本"] for m in raw_metrics]
    ratios = [m["最新期间原材料成本占比"] for m in raw_metrics]
    growths = [m["原材料成本增长率"] for m in raw_metrics]
    stucks = [m["360天以上积压金额_万元"] for m in raw_metrics]

    n_scales = minmax_normalize(scales)
    n_ratios = minmax_normalize(ratios)
    n_growths = minmax_normalize(growths)
    n_stucks = minmax_normalize(stucks)

    for i, m in enumerate(raw_metrics):
        m["规模标准化得分"] = round(n_scales[i], 4)
        m["占比标准化得分"] = round(n_ratios[i], 4)
        m["增长率标准化得分"] = round(n_growths[i], 4)
        m["积压标准化得分"] = round(n_stucks[i], 4)
        m["综合异常得分"] = round(
            n_scales[i] * BU_WEIGHTS["规模"]
            + n_ratios[i] * BU_WEIGHTS["占比"]
            + n_growths[i] * BU_WEIGHTS["增长率"]
            + n_stucks[i] * BU_WEIGHTS["积压"],
            4
        )

    # Rank
    raw_metrics.sort(key=lambda x: -x["综合异常得分"])
    for rank, m in enumerate(raw_metrics, 1):
        m["排名"] = rank

    return raw_metrics


# ── Composite scoring: Product Lines ─────────────────────────────
def score_product_lines(sheet1_rows, sheet2_rows):
    """Composite anomaly scoring for product lines."""
    pl_q, pl_bu = get_quarterly_by_pl(sheet1_rows)

    # Compute risk value inline (before analyze_materials is called)
    for m in sheet2_rows:
        m["风险值"] = round(m["库龄天数"] * m["原材料成本"] / 10000, 2)

    # High-risk material amounts by PL
    pl_risk_mat = defaultdict(float)
    for m in sheet2_rows:
        pl_risk_mat[m["产品线"]] += m["风险值"]

    pl_names = sorted(pl_q.keys())
    raw_metrics = []
    for pl in pl_names:
        latest = pl_q[pl].get(LATEST_QUARTER, {"总库存成本": 0, "原材料成本": 0})
        prev = pl_q[pl].get(PREV_QUARTER, {"总库存成本": 0, "原材料成本": 0})
        total = latest["总库存成本"]
        raw = latest["原材料成本"]
        ratio = round(raw / total * 100, 2) if total > 0 else 0
        growth = round((raw - prev["原材料成本"]) / prev["原材料成本"] * 100, 2) if prev["原材料成本"] > 0 else 0
        risk_amt = round(pl_risk_mat.get(pl, 0), 2)
        raw_metrics.append({
            "产品线": pl,
            "所属业务单元": pl_bu.get(pl, ""),
            "最新期间总库存成本": round(total, 2),
            "最新期间原材料成本": round(raw, 2),
            "最新期间原材料成本占比": ratio,
            "期初原材料成本": round(prev["原材料成本"], 2),
            "期末原材料成本": round(raw, 2),
            "原材料成本增长率": growth,
            "高风险物料金额": risk_amt,
        })

    scales = [m["最新期间原材料成本"] for m in raw_metrics]
    ratios = [m["最新期间原材料成本占比"] for m in raw_metrics]
    growths = [m["原材料成本增长率"] for m in raw_metrics]
    risk_amts = [m["高风险物料金额"] for m in raw_metrics]

    n_scales = minmax_normalize(scales)
    n_ratios = minmax_normalize(ratios)
    n_growths = minmax_normalize(growths)
    n_risks = minmax_normalize(risk_amts)

    for i, m in enumerate(raw_metrics):
        m["规模标准化得分"] = round(n_scales[i], 4)
        m["占比标准化得分"] = round(n_ratios[i], 4)
        m["增长率标准化得分"] = round(n_growths[i], 4)
        m["高风险物料标准化得分"] = round(n_risks[i], 4)
        m["综合异常得分"] = round(
            n_scales[i] * PL_WEIGHTS["规模"]
            + n_ratios[i] * PL_WEIGHTS["占比"]
            + n_growths[i] * PL_WEIGHTS["增长率"]
            + n_risks[i] * PL_WEIGHTS["高风险物料"],
            4
        )

    raw_metrics.sort(key=lambda x: -x["综合异常得分"])
    for rank, m in enumerate(raw_metrics, 1):
        m["排名"] = rank

    return raw_metrics


# ── Build product line quarterly trends ──────────────────────────
def build_pl_trends(sheet1_rows, top_n=5):
    """Build quarterly trend series for top product lines by composite score."""
    pl_scores = score_product_lines(sheet1_rows, [])  # no material data needed for ranking
    # Actually we need the full scored list — but for trend visualization
    # we use top by raw material cost for comprehensiveness
    pl_cost = defaultdict(float)
    for r in sheet1_rows:
        pl_cost[r["产品线"]] += r["原材料成本"]
    top_pls = sorted(pl_cost.keys(), key=lambda x: -pl_cost[x])[:top_n]

    result = {}
    for pl in top_pls:
        qmap = defaultdict(lambda: {"原材料成本": 0.0, "总库存成本": 0.0})
        for r in sheet1_rows:
            if r["产品线"] == pl:
                q = r["季度"]
                qmap[q]["原材料成本"] += r["原材料成本"]
                qmap[q]["总库存成本"] += r["总库存成本"]
        qs = sorted(qmap.keys())
        result[pl] = []
        for q in qs:
            d = qmap[q]
            result[pl].append({
                "季度": q,
                "原材料成本": round(d["原材料成本"], 2),
                "总库存成本": round(d["总库存成本"], 2),
                "原材料成本占比": round(d["原材料成本"] / d["总库存成本"] * 100, 2) if d["总库存成本"] > 0 else 0,
            })
    return result


# ── Build KPI ────────────────────────────────────────────────────
def build_kpis(structure_summary, material_analysis, bu_scores, pl_scores, quarterly_trend):
    """Compute all KPI metrics."""
    latest_q = structure_summary[-1]
    prev_q = structure_summary[-2]

    total_inv = round(latest_q["总库存成本"] / 10000, 2)
    raw_cost = round(latest_q["原材料成本"] / 10000, 2)
    raw_ratio = latest_q["原材料成本占比"]
    raw_growth = round((latest_q["原材料成本"] - prev_q["原材料成本"]) / prev_q["原材料成本"] * 100, 2)
    stuck_360 = round(material_analysis["积压金额_360天以上"] / 10000, 2)
    high_risk_count = len(material_analysis["超过360天物料"])

    top_bu = bu_scores[0]
    top_pl = pl_scores[0]

    # Trend data for KPI section
    trend_list = []
    for q in quarterly_trend:
        trend_list.append({
            "季度": q["季度"],
            "总库存成本_万元": round(q["总库存成本"] / 10000, 2),
            "原材料成本_万元": round(q["原材料成本"] / 10000, 2),
            "半成品成本_万元": round(q["半成品成本"] / 10000, 2),
            "产成品成本_万元": round(q["产成品成本"] / 10000, 2),
            "原材料成本占比": q["原材料成本占比"],
        })

    return {
        "总库存成本_万元": total_inv,
        "原材料成本_万元": raw_cost,
        "原材料成本占比": raw_ratio,
        "原材料成本环比增长": raw_growth,
        "积压金额_360天以上_万元": stuck_360,
        "高风险物料数量": high_risk_count,
        "异常业务单元": top_bu["业务单元"],
        "异常产品线": top_pl["产品线"],
        "异常业务单元得分": top_bu["综合异常得分"],
        "异常产品线得分": top_pl["综合异常得分"],
        "数据期间": "2020年-2025年Q1",
        "季度趋势": trend_list,
    }


# ── Build debug JSON ─────────────────────────────────────────────
def build_debug(sheet1_info, sheet2_info, sheet3_info,
                structure_summary, quarterly_trend,
                bu_scores, pl_scores, material_analysis, kpis):
    """Build comprehensive debug output for manual verification."""
    return {
        "_说明": "本文件用于人工核对每个指标的计算过程。所有金额单位为原始单位（元），万元另有标注。",
        "数据来源": "比赛数据.xlsx",
        "工作表信息": {
            "工作表1": sheet1_info,
            "工作表2": sheet2_info,
            "工作表3": sheet3_info,
        },
        "KPI计算过程": {
            "最新期间": LATEST_QUARTER,
            "上一期间": PREV_QUARTER,
            "总库存成本_万元": kpis["总库存成本_万元"],
            "总库存成本_计算": f"Sheet3 {LATEST_QUARTER} 原材料+半成品+产成品 = {structure_summary[-1]['总库存成本']:.2f} 元 = {kpis['总库存成本_万元']:.2f} 万元",
            "原材料成本_万元": kpis["原材料成本_万元"],
            "原材料成本_计算": f"Sheet3 {LATEST_QUARTER} 原材料成本 = {structure_summary[-1]['原材料成本']:.2f} 元 = {kpis['原材料成本_万元']:.2f} 万元",
            "原材料成本占比": kpis["原材料成本占比"],
            "原材料成本占比_计算": f"{kpis['原材料成本_万元']:.2f} / {kpis['总库存成本_万元']:.2f} × 100% = {kpis['原材料成本占比']:.2f}%",
            "原材料成本环比增长": kpis["原材料成本环比增长"],
            "原材料成本环比增长_计算": f"({LATEST_QUARTER} {structure_summary[-1]['原材料成本']:.2f} - {PREV_QUARTER} {structure_summary[-2]['原材料成本']:.2f}) / {structure_summary[-2]['原材料成本']:.2f} × 100% = {kpis['原材料成本环比增长']:.2f}%",
            "360天以上积压金额_万元": kpis["积压金额_360天以上_万元"],
            "360天以上积压金额_计算": f"Sheet2 物料明细中库龄>360天的物料原材料成本合计 = {material_analysis['积压金额_360天以上']:.2f} 元 = {kpis['积压金额_360天以上_万元']:.2f} 万元",
            "高风险物料数量": kpis["高风险物料数量"],
            "异常识别方法": "采用综合贡献度评分：业务单元=规模×0.40+占比×0.25+增长率×0.20+积压×0.15；产品线=规模×0.35+占比×0.20+增长率×0.20+高风险物料×0.25。所有子项均经min-max标准化。",
            "异常业务单元": kpis["异常业务单元"],
            "异常业务单元得分": kpis["异常业务单元得分"],
            "异常产品线": kpis["异常产品线"],
            "异常产品线得分": kpis["异常产品线得分"],
        },
        "业务单元综合排名": bu_scores,
        "业务单元评分权重": BU_WEIGHTS,
        "产品线综合排名": pl_scores,
        "产品线评分权重": PL_WEIGHTS,
        "物料风险排名TOP10": material_analysis["高风险物料TOP10"],
        "物料评分权重": MAT_WEIGHTS,
        "物料汇总": material_analysis["物料汇总"],
        "库龄分布": material_analysis["库龄分布"],
        "季度成本结构_万元": [
            {
                "季度": s["季度"],
                "原材料成本_万元": round(s["原材料成本"] / 10000, 2),
                "半成品成本_万元": round(s["半成品成本"] / 10000, 2),
                "产成品成本_万元": round(s["产成品成本"] / 10000, 2),
                "总库存成本_万元": round(s["总库存成本"] / 10000, 2),
                "原材料成本占比": s["原材料成本占比"],
            }
            for s in structure_summary
        ],
    }


# ── Main ─────────────────────────────────────────────────────────
def main():
    print("Loading Excel...")
    wb, xlsx_path = load_workbook()
    sheets = wb.sheetnames
    print(f"Workbook: {xlsx_path}")
    print(f"Sheets ({len(sheets)}): {sheets[0]}, {sheets[1]}, {sheets[2]}")

    # Parse
    print("Parsing sheet 1...")
    s1_rows, s1_info = parse_sheet1(wb[sheets[0]])
    print("Parsing sheet 2...")
    s2_rows, s2_info = parse_sheet2(wb[sheets[1]])
    print("Parsing sheet 3...")
    s3_rows, s3_info = parse_sheet3(wb[sheets[2]])

    # Build
    print("Building structure summary...")
    structure_summary = build_structure_summary(s3_rows)
    print("Aggregating quarterly trends...")
    quarterly_trend = aggregate_by_quarter(s1_rows)

    print("Scoring business units...")
    bu_scores = score_business_units(s1_rows, s2_rows, structure_summary)
    print("Scoring product lines...")
    pl_scores = score_product_lines(s1_rows, s2_rows)
    print("Analyzing materials...")
    mat_analysis = analyze_materials(s2_rows)
    print("Building PL trends...")
    pl_trends = build_pl_trends(s1_rows)

    print("Building KPIs...")
    kpis = build_kpis(structure_summary, mat_analysis, bu_scores, pl_scores, quarterly_trend)

    # Build bubble data for BU chart
    bu_bubble = []
    bu_q = get_quarterly_by_bu(s1_rows)
    for bu in sorted(bu_q.keys()):
        latest = bu_q[bu].get(LATEST_QUARTER, {"总库存成本": 0, "原材料成本": 0})
        prev = bu_q[bu].get(PREV_QUARTER, {"原材料成本": 0})
        cost = latest["原材料成本"]
        growth = round((cost - prev["原材料成本"]) / prev["原材料成本"] * 100, 2) if prev["原材料成本"] > 0 else 0
        ratio = round(cost / latest["总库存成本"] * 100, 2) if latest["总库存成本"] > 0 else 0
        bu_bubble.append({
            "业务单元": bu,
            "原材料成本_万元": round(cost / 10000, 2),
            "增长率": growth,
            "原材料成本占比": ratio,
            "综合异常得分": next((b["综合异常得分"] for b in bu_scores if b["业务单元"] == bu), 0),
        })

    # Build chart-ready data
    structure_for_chart = []
    for s in structure_summary:
        structure_for_chart.append({
            "季度": s["季度"],
            "原材料成本_万元": round(s["原材料成本"] / 10000, 2),
            "半成品成本_万元": round(s["半成品成本"] / 10000, 2),
            "产成品成本_万元": round(s["产成品成本"] / 10000, 2),
            "总库存成本_万元": round(s["总库存成本"] / 10000, 2),
            "原材料成本占比": s["原材料成本占比"],
        })

    bu_for_chart = []
    for b in bu_scores:
        bu_for_chart.append({
            "业务单元": b["业务单元"],
            "总库存成本_万元": round(b["最新期间总库存成本"] / 10000, 2),
            "原材料成本_万元": round(b["最新期间原材料成本"] / 10000, 2),
            "原材料成本占比": b["最新期间原材料成本占比"],
            "原材料成本增长率": b["原材料成本增长率"],
            "360天以上积压金额_万元": b["360天以上积压金额_万元"],
            "综合异常得分": b["综合异常得分"],
            "排名": b["排名"],
        })

    pl_for_chart = []
    for p in pl_scores:
        pl_for_chart.append({
            "产品线": p["产品线"],
            "所属业务单元": p["所属业务单元"],
            "总库存成本_万元": round(p["最新期间总库存成本"] / 10000, 2),
            "原材料成本_万元": round(p["最新期间原材料成本"] / 10000, 2),
            "原材料成本占比": p["最新期间原材料成本占比"],
            "原材料成本增长率": p["原材料成本增长率"],
            "高风险物料金额": p["高风险物料金额"],
            "综合异常得分": p["综合异常得分"],
            "排名": p["排名"],
        })

    # Assemble dashboard data
    dashboard_data = {
        "kpi": kpis,
        "库存成本结构": structure_for_chart,
        "季度趋势": quarterly_trend,
        "业务单元分析": bu_for_chart,
        "业务单元气泡图": bu_bubble,
        "产品线分析": pl_for_chart,
        "产品线趋势": pl_trends,
        "物料分析": {
            "物料明细": mat_analysis["物料明细"],
            "超过360天物料": mat_analysis["超过360天物料"],
            "积压金额_360天以上_万元": round(mat_analysis["积压金额_360天以上"] / 10000, 2),
            "库龄分布": mat_analysis["库龄分布"],
            "高风险物料TOP10": mat_analysis["高风险物料TOP10"],
            "物料汇总": mat_analysis["物料汇总"],
        },
        "综合评分": {
            "业务单元排名": [{"排名": b["排名"], "业务单元": b["业务单元"], "综合异常得分": b["综合异常得分"]} for b in bu_scores],
            "产品线排名": [{"排名": p["排名"], "产品线": p["产品线"], "综合异常得分": p["综合异常得分"]} for p in pl_scores],
        },
        "数据来源": "比赛数据.xlsx",
        "数据期间": "2020年-2025年Q1",
    }

    # Assemble debug data
    debug_data = build_debug(
        s1_info, s2_info, s3_info,
        structure_summary, quarterly_trend,
        bu_scores, pl_scores, mat_analysis, kpis
    )

    # Write outputs
    script_path = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_path)
    data_dir = os.path.join(project_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    dash_path = os.path.join(data_dir, "dashboard-data.json")
    with open(dash_path, "w", encoding="utf-8") as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    debug_path = os.path.join(data_dir, "analysis-debug.json")
    with open(debug_path, "w", encoding="utf-8") as f:
        json.dump(debug_data, f, ensure_ascii=False, indent=2)

    # Print summary
    print(f"\n{'='*60}")
    print(f"dashboard-data.json -> {dash_path}")
    print(f"analysis-debug.json -> {debug_path}")
    print(f"\nKPI Summary:")
    print(f"  Total inventory cost: {kpis['总库存成本_万元']} wan yuan")
    print(f"  Raw material cost:    {kpis['原材料成本_万元']} wan yuan")
    print(f"  Raw material ratio:   {kpis['原材料成本占比']}%")
    print(f"  Raw material growth:  {kpis['原材料成本环比增长']}% (QoQ)")
    print(f"  Stuck >360d:          {kpis['积压金额_360天以上_万元']} wan yuan")
    print(f"  High-risk materials:  {kpis['高风险物料数量']}")
    print(f"\nBusiness Unit Composite Ranking:")
    for b in bu_scores:
        print(f"  #{b['排名']} {b['业务单元']}: score={b['综合异常得分']:.4f} (规模={b['规模标准化得分']:.4f}, 占比={b['占比标准化得分']:.4f}, 增长={b['增长率标准化得分']:.4f}, 积压={b['积压标准化得分']:.4f})")
    print(f"\nProduct Line Composite Ranking (Top 5):")
    for p in pl_scores[:5]:
        print(f"  #{p['排名']} {p['产品线']} ({p['所属业务单元']}): score={p['综合异常得分']:.4f}")
    print(f"\nMaterial Risk TOP5:")
    for i, m in enumerate(mat_analysis['高风险物料TOP10'][:5]):
        print(f"  #{i+1} {m['物料名称']}: risk_score={m['风险得分']:.4f}, aging={m['库龄天数']}d, amount={m['原材料成本']:.2f}")
    print(f"\nAbnormal BU:  {kpis['异常业务单元']} (score={kpis['异常业务单元得分']:.4f})")
    print(f"Abnormal PL:  {kpis['异常产品线']} (score={kpis['异常产品线得分']:.4f})")
    print(f"\nDone. Both JSON files written successfully.")


if __name__ == "__main__":
    main()
